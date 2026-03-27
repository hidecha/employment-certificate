"""
就労証明書Excelフォーマット解析スクリプト

Usage:
    python analyze_shuroushomei.py <workbook_path> [--json]

Example:
    python analyze_shuroushomei.py shuroushomei.xlsx --json
"""

import sys
import re
import json
import logging
import openpyxl
from collections import Counter
from openpyxl.utils import get_column_letter
from pathlib import Path

logger = logging.getLogger(__name__)

EXPECTED_TEXT_COUNT = 125
EXPECTED_CHECKBOX_COUNT = 80


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def get_merge_master(ws, row: int, col: int) -> tuple[int, int]:
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            return (merged_range.min_row, merged_range.min_col)
    return (row, col)


def cell_addr(row: int, col: int) -> str:
    return f'{get_column_letter(col)}{row}'


def find_cells_by_text(ws, text: str, start_row: int = 1, end_row: int = 60) -> list[tuple[int, int, str]]:
    results = []
    for row in range(start_row, min(end_row + 1, ws.max_row + 1)):
        for cell in ws[row]:
            if cell.value and text in str(cell.value):
                results.append((cell.row, cell.column, str(cell.value)))
    return results


def find_input_before_label(ws, label_row: int, label_col: int) -> str | None:
    for col in range(label_col - 1, max(0, label_col - 8), -1):
        master_row, master_col = get_merge_master(ws, label_row, col)
        cell_val = ws.cell(row=master_row, column=master_col).value
        if cell_val and str(cell_val).strip() not in ('', '西暦'):
            break
        if (master_row, master_col) != (label_row, col):
            return cell_addr(master_row, master_col)
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= label_row <= mr.max_row and mr.min_col == col:
                return cell_addr(mr.min_row, mr.min_col)
    fallback_col = label_col - 1
    if fallback_col > 0:
        mr, mc = get_merge_master(ws, label_row, fallback_col)
        return cell_addr(mr, mc)
    return None


def find_merged_input_at(ws, row: int, col: int) -> str:
    mr, mc = get_merge_master(ws, row, col)
    if (mr, mc) != (row, col):
        return cell_addr(mr, mc)
    for offset in range(0, 3):
        check_col = col + offset
        if offset > 0:
            val = ws.cell(row=row, column=check_col).value
            if val and str(val).strip() != '':
                break
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col == check_col):
                return cell_addr(merged_range.min_row, merged_range.min_col)
    return cell_addr(row, col)


def find_input_in_range(ws, row: int, start_col: int, end_col: int) -> str | None:
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                start_col <= merged_range.min_col < end_col):
            return cell_addr(merged_range.min_row, merged_range.min_col)
    for col in range(start_col, end_col):
        val = ws.cell(row=row, column=col).value
        if not val or str(val).strip() == '':
            return cell_addr(row, col)
    return None


def find_input_after_dash(ws, row: int, dash_col: int) -> str | None:
    next_col = dash_col + 1
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= dash_col <= mr.max_col:
            next_col = mr.max_col + 1
            break
    return find_merged_input_at(ws, row, next_col)


def find_right_panel_input(ws, row: int, label_col: int) -> str:
    input_col = label_col + 1
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= label_col <= mr.max_col:
            input_col = mr.max_col + 1
            break
    return find_merged_input_at(ws, row, input_col)


def find_phone_fields(ws, row: int, label_end: int, prefix: str) -> dict[str, str]:
    mapping = {}
    dashes = []
    for c in ws[row]:
        if c.column > label_end and c.value and str(c.value).strip() in ('―', '-', '－'):
            dashes.append(c.column)
    if len(dashes) >= 2:
        mapping[f'{prefix}1'] = find_input_before_label(ws, row, dashes[0])
        mapping[f'{prefix}2'] = find_input_after_dash(ws, row, dashes[0])
        mapping[f'{prefix}3'] = find_input_after_dash(ws, row, dashes[1])
    elif len(dashes) == 1:
        mapping[f'{prefix}1'] = find_input_before_label(ws, row, dashes[0])
        mapping[f'{prefix}2'] = find_input_after_dash(ws, row, dashes[0])
    else:
        mapping[f'{prefix}1'] = find_merged_input_at(ws, row, label_end + 1)
    return mapping


def analyze_ymd_row(ws, row: int, start_col: int = 1) -> dict[str, str | None]:
    """年・月・日ラベルを探し、その直前の入力セルを返す"""
    year_labels = [(c.row, c.column) for c in ws[row]
                   if c.column >= start_col and c.value and str(c.value).strip() == '年']
    month_labels = [(c.row, c.column) for c in ws[row]
                    if c.column >= start_col and c.value and str(c.value).strip() == '月']
    day_labels = [(c.row, c.column) for c in ws[row]
                  if c.column >= start_col and c.value and str(c.value).strip() == '日']
    result = {}
    if year_labels:
        result['年'] = find_input_before_label(ws, row, year_labels[0][1])
    if month_labels:
        result['月'] = find_input_before_label(ws, row, month_labels[0][1])
    if day_labels:
        result['日'] = find_input_before_label(ws, row, day_labels[0][1])
    return result


def analyze_period_row(ws, row: int, start_col: int = 1) -> dict[str, str | None]:
    """年月日～年月日パターンの行を解析"""
    tilde_col = 999
    for c in ws[row]:
        if c.column >= start_col and c.value and '～' in str(c.value):
            tilde_col = c.column
            break

    year_labels = [(c.row, c.column) for c in ws[row]
                   if c.column >= start_col and c.value and str(c.value).strip() == '年']
    month_labels = [(c.row, c.column) for c in ws[row]
                    if c.column >= start_col and c.value and str(c.value).strip() == '月']
    day_labels = [(c.row, c.column) for c in ws[row]
                  if c.column >= start_col and c.value and str(c.value).strip() == '日']

    result = {}
    for yl in year_labels:
        if yl[1] < tilde_col:
            result['開始・年'] = find_input_before_label(ws, row, yl[1])
            break
    for ml in month_labels:
        if ml[1] < tilde_col:
            result['開始・月'] = find_input_before_label(ws, row, ml[1])
            break
    for dl in day_labels:
        if dl[1] < tilde_col:
            result['開始・日'] = find_input_before_label(ws, row, dl[1])
            break
    for yl in year_labels:
        if yl[1] > tilde_col:
            result['終了・年'] = find_input_before_label(ws, row, yl[1])
            break
    for ml in month_labels:
        if ml[1] > tilde_col:
            result['終了・月'] = find_input_before_label(ws, row, ml[1])
            break
    for dl in day_labels:
        if dl[1] > tilde_col:
            result['終了・日'] = find_input_before_label(ws, row, dl[1])
            break
    return result


def analyze_time_row(ws, row: int) -> dict[str, str | None]:
    """時・分・～・時・分・（うち休憩時間・分）パターンの行を解析"""
    hour_cells = [(c.row, c.column) for c in ws[row]
                  if c.value and str(c.value).strip() == '時']
    min_cells = [(c.row, c.column) for c in ws[row]
                 if c.value and str(c.value).strip() == '分'
                 and not any('休憩' in str(ws.cell(row=row, column=cc).value or '')
                             for cc in range(max(1, c.column - 5), c.column))]
    rest_cells = [(c.row, c.column) for c in ws[row]
                  if c.value and '休憩' in str(c.value)]

    start_hour = start_min = end_hour = end_min = rest_time = None
    if len(hour_cells) >= 2:
        start_hour = find_input_before_label(ws, row, hour_cells[0][1])
        end_hour = find_input_before_label(ws, row, hour_cells[1][1])
    if len(min_cells) >= 2:
        start_min = find_input_before_label(ws, row, min_cells[0][1])
        end_min = find_input_before_label(ws, row, min_cells[1][1])
    elif len(min_cells) == 1:
        start_min = find_input_before_label(ws, row, min_cells[0][1])
    if rest_cells:
        for rc in rest_cells:
            paren = [(c.row, c.column) for c in ws[row]
                     if c.value and '分）' in str(c.value) and c.column > rc[1]]
            if paren:
                rest_time = find_input_before_label(ws, row, paren[0][1])
                break
    result = {}
    if start_hour:
        result['開始・時'] = start_hour
    if start_min:
        result['開始・分'] = start_min
    if end_hour:
        result['終了・時'] = end_hour
    if end_min:
        result['終了・分'] = end_min
    if rest_time:
        result['休憩時間'] = rest_time
    return result


def _find_other_text_input(ws, search_text: str, row_start: int, row_end: int) -> str | None:
    """「その他（ ）」パターンのテキスト入力セルを探す。"""
    other_cells = find_cells_by_text(ws, f'その他（', row_start, row_end)
    if not other_cells:
        other_cells = find_cells_by_text(ws, f'その他(', row_start, row_end)
    if not other_cells:
        return None
    or_row, or_col = other_cells[0][0], other_cells[0][1]
    paren_close = [(c.row, c.column) for c in ws[or_row]
                   if c.value and str(c.value).strip() in ('）', ')') and c.column > or_col]
    if paren_close:
        return find_input_in_range(ws, or_row, or_col + 1, paren_close[0][1])
    for mr in ws.merged_cells.ranges:
        if mr.min_row == or_row and mr.min_col > or_col:
            return cell_addr(mr.min_row, mr.min_col)
    return None


def _get_label_end_col(ws, row: int, col: int) -> int:
    """ラベルセルの結合範囲の右端列を返す。"""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return mr.max_col
    return col


# ---------------------------------------------------------------------------
# Municipality detection from content
# ---------------------------------------------------------------------------

def _select_form_sheet(wb) -> 'openpyxl.worksheet.worksheet.Worksheet':
    """就労証明書のフォームシートを選択する。

    括弧付きプレフィックス（案内・記載例など）のないシートを優先し、
    見つからない場合は先頭シートにフォールバックする。
    """
    _SKIP_SHEETS = {'プルダウンリスト', '記載要領'}
    candidates = []
    for ws in wb.worksheets:
        title = ws.title.strip()
        if title in _SKIP_SHEETS:
            continue
        # 全角括弧プレフィックスを除外したシートを優先
        if not re.match(r'^[（(].+[）)]', title):
            candidates.append(ws)
    if candidates:
        return candidates[0]
    # フォールバック: 先頭シート
    return wb.worksheets[0]


def extract_municipality_from_content(ws) -> str:
    """ワークシートの内容から自治体名を推測する"""
    for row in range(1, 10):
        for cell in ws[row]:
            if not cell.value:
                continue
            val = str(cell.value).strip()
            for suffix in ['市長', '区長', '町長', '村長']:
                if suffix in val:
                    idx = val.index(suffix)
                    return val[:idx + len(suffix) - 1].strip()

    for row in range(1, 6):
        for cell in ws[row]:
            if cell.value and str(cell.value).strip() == '宛':
                for col in range(cell.column - 1, 0, -1):
                    left_val = ws.cell(row=row, column=col).value
                    if left_val and str(left_val).strip():
                        candidate = str(left_val).strip()
                        if any(k in candidate for k in ('所長', 'センター', '福祉事務')):
                            break
                        return candidate

    municipality_pattern = re.compile(r'(.+?(?:都|道|府|県))?(.+?(?:市|区|町|村))')
    for row in range(1, 8):
        for cell in ws[row]:
            if cell.value:
                val = str(cell.value)
                if '就労' in val or any(k in val for k in ('所長', 'センター', '福祉事務')):
                    continue
                m = municipality_pattern.search(val)
                if m:
                    return m.group(0).strip()

    return "不明"


# ---------------------------------------------------------------------------
# Section-specific text field parsers
# ---------------------------------------------------------------------------

def _parse_addressee(ws, mapping: dict) -> None:
    """宛先セクションを解析する。"""
    for row in range(1, 12):
        for cell in ws[row]:
            if cell.value:
                val = str(cell.value)
                if any(k in val for k in ('市長', '区長', '町長', '村長', '福祉事務所長', 'センター所長')):
                    mr, mc = get_merge_master(ws, cell.row, cell.column)
                    mapping['宛先'] = cell_addr(mr, mc)
                    return
    # フォールバック: 「宛」ラベルの左側の結合セル
    for row in range(1, 12):
        for cell in ws[row]:
            if cell.value and str(cell.value).strip() == '宛':
                for mr in ws.merged_cells.ranges:
                    if mr.min_row <= row <= mr.max_row and mr.max_col < cell.column:
                        mapping['宛先'] = cell_addr(mr.min_row, mr.min_col)
                        return


def _parse_certification_date(ws, mapping: dict) -> None:
    """証明日（年月日）セクションを解析する。"""
    cert_date_cells = find_cells_by_text(ws, '証明日', 1, 10)
    cert_date_cells = sorted(cert_date_cells, key=lambda c: len(c[2]))
    cert_date_cells = [c for c in cert_date_cells if len(c[2].strip()) <= 6] or cert_date_cells
    if not cert_date_cells:
        return
    cert_row = cert_date_cells[0][0]
    logger.debug(f"  証明日ラベル: row={cert_row}, text={cert_date_cells[0][2]!r}")
    ymd = analyze_ymd_row(ws, cert_row, cert_date_cells[0][1])
    for k in ('年', '月', '日'):
        if k in ymd:
            mapping[f'証明日・{k}'] = ymd[k]


def _parse_office_info(ws, mapping: dict) -> None:
    """右上パネル（事業所情報）を解析する。"""
    for row in range(3, 12):
        for cell in ws[row]:
            if not cell.value or cell.column < 20:
                continue
            val = str(cell.value).strip()

            if val == '事業所名':
                mapping['事業所名'] = find_right_panel_input(ws, row, cell.column)
            elif val == '代表者名':
                mapping['代表者名'] = find_right_panel_input(ws, row, cell.column)
            elif val == '所在地':
                mapping['所在地'] = find_right_panel_input(ws, row, cell.column)
            elif val == '電話番号':
                label_end = _get_label_end_col(ws, row, cell.column)
                mapping.update(find_phone_fields(ws, row, label_end, '電話番号'))
            elif val == '担当者名':
                mapping['担当者名'] = find_right_panel_input(ws, row, cell.column)
            elif val == '記載者連絡先':
                label_end = _get_label_end_col(ws, row, cell.column)
                mapping.update(find_phone_fields(ws, row, label_end, '記載者連絡先'))


def _parse_industry(ws, mapping: dict) -> None:
    """No.1 業種セクションを解析する。"""
    result = _find_other_text_input(ws, 'その他', 13, 18)
    if result:
        mapping['業種・その他'] = result


def _parse_personal_info(ws, mapping: dict) -> None:
    """No.2 フリガナ・本人氏名・生年月日セクションを解析する。"""
    furigana_cells = find_cells_by_text(ws, 'フリガナ', 14, 25)
    if furigana_cells:
        fr = furigana_cells[0][0]
        for mr in ws.merged_cells.ranges:
            if mr.min_row == fr and mr.min_col >= 8 and (mr.max_col - mr.min_col) >= 5:
                mapping['フリガナ'] = cell_addr(mr.min_row, mr.min_col)
                break

    name_cells = find_cells_by_text(ws, '本人氏名', 14, 25)
    if name_cells:
        nr = name_cells[0][0]
        for mr in ws.merged_cells.ranges:
            if mr.min_row == nr and mr.min_col >= 8 and (mr.max_col - mr.min_col) >= 5:
                mapping['本人氏名'] = cell_addr(mr.min_row, mr.min_col)
                break

    birth_cells = find_cells_by_text(ws, '生年', 14, 25)
    if birth_cells:
        br = birth_cells[0][0]
        ymd = analyze_ymd_row(ws, br, birth_cells[0][1])
        for k in ('年', '月', '日'):
            if k in ymd:
                mapping[f'生年月日・{k}'] = ymd[k]


def _parse_employment_period(ws, mapping: dict) -> None:
    """No.3 雇用(予定)期間等セクションを解析する。"""
    employ_cells = find_cells_by_text(ws, '雇用', 15, 25)
    employ_cells = [c for c in employ_cells if '期間' in c[2]]
    if employ_cells:
        er = employ_cells[0][0]
        period = analyze_period_row(ws, er, employ_cells[0][1])
        for k, v in period.items():
            mapping[f'雇用期間・{k}'] = v


def _parse_work_office(ws, mapping: dict) -> None:
    """No.4 本人就労先事業所セクションを解析する。"""
    office_cells = find_cells_by_text(ws, '本人就労先', 18, 25)
    if not office_cells:
        return
    orow = office_cells[0][0]
    for search_row in range(orow, orow + 3):
        for cell in ws[search_row]:
            if cell.value and str(cell.value).strip() == '名称':
                mapping['事業所名称'] = find_right_panel_input(ws, search_row, cell.column)
            elif cell.value and str(cell.value).strip() == '住所':
                mapping['事業所住所'] = find_right_panel_input(ws, search_row, cell.column)


def _parse_employment_type(ws, mapping: dict) -> None:
    """No.5 雇用の形態セクションを解析する。"""
    employ_type_others = find_cells_by_text(ws, 'その他(', 22, 26)
    if not employ_type_others:
        employ_type_others = find_cells_by_text(ws, 'その他（', 22, 26)
    if not employ_type_others:
        return
    for eto in employ_type_others:
        paren_close = [(c.row, c.column) for c in ws[eto[0]]
                       if c.value and str(c.value).strip() in ('）', ')') and c.column > eto[1]]
        if paren_close:
            mapping['雇用形態・その他'] = find_input_in_range(ws, eto[0], eto[1] + 1, paren_close[0][1])
            break
        for mr in ws.merged_cells.ranges:
            if mr.min_row == eto[0] and mr.min_col > eto[1]:
                mapping['雇用形態・その他'] = cell_addr(mr.min_row, mr.min_col)
                break


def _parse_fixed_work_time(ws, mapping: dict, work_time_cells: list) -> None:
    """No.6 就労時間（固定就労の場合）を解析する。"""
    fixed_cells = [c for c in work_time_cells if '固定' in c[2]]
    if not fixed_cells:
        return
    wt_row = fixed_cells[0][0]

    # 月間合計
    monthly_label = None
    for r in range(wt_row, wt_row + 3):
        for c in ws[r]:
            if c.value and str(c.value).strip() == '月間':
                monthly_label = (r, c.column)
                break
        if monthly_label:
            break

    if monthly_label:
        ml_row = monthly_label[0]
        hour_labels = [(c.row, c.column) for c in ws[ml_row]
                       if c.value and str(c.value).strip() == '時間' and c.column > monthly_label[1]]
        min_labels = [(c.row, c.column) for c in ws[ml_row]
                      if c.value and str(c.value).strip() == '分' and c.column > monthly_label[1]]
        rest_labels = [(c.row, c.column) for c in ws[ml_row]
                       if c.value and '休憩' in str(c.value) and c.column > monthly_label[1]]

        if hour_labels:
            mapping['就労時間・月間・時間'] = find_input_before_label(ws, ml_row, hour_labels[0][1])
        if min_labels:
            mapping['就労時間・月間・分'] = find_input_before_label(ws, ml_row, min_labels[0][1])
        if rest_labels:
            rest_row = rest_labels[0][0]
            rest_col = rest_labels[0][1]
            paren = [(c.row, c.column) for c in ws[rest_row]
                     if c.value and '分）' in str(c.value) and c.column > rest_col]
            if paren:
                mapping['就労時間・月間・休憩時間'] = find_input_before_label(ws, rest_row, paren[0][1])

    # 一月/一週当たり就労日数
    monthly_days_cells = find_cells_by_text(ws, '一月当たり', wt_row, wt_row + 6)
    if not monthly_days_cells:
        monthly_days_cells = find_cells_by_text(ws, '1月当たり', wt_row, wt_row + 6)
    if monthly_days_cells:
        md_row = monthly_days_cells[0][0]
        day_labels_on_row = [(c.row, c.column) for c in ws[md_row]
                             if c.value and str(c.value).strip() == '日']
        if day_labels_on_row:
            mapping['就労時間・一月当たりの就労日数'] = find_input_before_label(ws, md_row, day_labels_on_row[0][1])

    weekly_cells = find_cells_by_text(ws, '一週当たり', wt_row, wt_row + 6)
    if not weekly_cells:
        weekly_cells = find_cells_by_text(ws, '1週当たり', wt_row, wt_row + 6)
    if weekly_cells:
        wk_row = weekly_cells[0][0]
        wk_col = weekly_cells[0][1]
        day_labels_weekly = [(c.row, c.column) for c in ws[wk_row]
                             if c.value and str(c.value).strip() == '日' and c.column > wk_col]
        if day_labels_weekly:
            mapping['就労時間・一週当たりの就労日数'] = find_input_before_label(ws, wk_row, day_labels_weekly[0][1])

    # 平日・土曜・日祝
    for day_label, prefix in [('平日', '就労時間・平日'), ('土曜', '就労時間・土曜'), ('日祝', '就労時間・日祝')]:
        day_cells = find_cells_by_text(ws, day_label, wt_row, wt_row + 10)
        if day_cells:
            tr = analyze_time_row(ws, day_cells[0][0])
            for k, v in tr.items():
                mapping[f'{prefix}・{k}'] = v


def _parse_irregular_work_time(ws, mapping: dict, work_time_cells: list) -> None:
    """No.6 就労時間（変則就労の場合）を解析する。"""
    irregular_cells = [c for c in work_time_cells if '変則' in c[2]]
    if not irregular_cells:
        return
    ir_row = irregular_cells[0][0]

    # 合計時間
    total_cells = find_cells_by_text(ws, '合計時間', ir_row, ir_row + 3)
    if total_cells:
        t_row = total_cells[0][0]
        hour_labels = [(c.row, c.column) for c in ws[t_row]
                       if c.value and str(c.value).strip() == '時間']
        min_labels = [(c.row, c.column) for c in ws[t_row]
                      if c.value and str(c.value).strip() == '分'
                      and not any('休憩' in str(ws.cell(row=t_row, column=cc).value or '')
                                  for cc in range(max(1, c.column - 5), c.column))]
        rest_labels = [(c.row, c.column) for c in ws[t_row]
                       if c.value and '休憩' in str(c.value)]
        if hour_labels:
            mapping['変則就労・合計時間・時間'] = find_input_before_label(ws, t_row, hour_labels[0][1])
        if min_labels:
            mapping['変則就労・合計時間・分'] = find_input_before_label(ws, t_row, min_labels[0][1])
        if rest_labels:
            paren = [(c.row, c.column) for c in ws[t_row]
                     if c.value and '分）' in str(c.value) and c.column > rest_labels[0][1]]
            if paren:
                mapping['変則就労・合計時間・休憩時間'] = find_input_before_label(ws, t_row, paren[0][1])

    # 就労日数
    days_cells = find_cells_by_text(ws, '就労日数', ir_row, ir_row + 4)
    if days_cells:
        d_row = days_cells[0][0]
        day_labels = [(c.row, c.column) for c in ws[d_row]
                      if c.value and str(c.value).strip() == '日']
        if day_labels:
            mapping['変則就労・就労日数'] = find_input_before_label(ws, d_row, day_labels[0][1])

    # 主な就労時間帯
    shift_cells = find_cells_by_text(ws, '主な就労時間帯', ir_row, ir_row + 5)
    if not shift_cells:
        shift_cells = find_cells_by_text(ws, 'シフト', ir_row, ir_row + 5)
    if shift_cells:
        tr = analyze_time_row(ws, shift_cells[0][0])
        for k, v in tr.items():
            mapping[f'変則就労・就労時間帯・{k}'] = v


def _parse_work_time(ws, mapping: dict) -> None:
    """No.6 就労時間セクション全体を解析する。"""
    work_time_cells = find_cells_by_text(ws, '就労時間', 22, 35)
    _parse_fixed_work_time(ws, mapping, work_time_cells)
    _parse_irregular_work_time(ws, mapping, work_time_cells)


def _parse_work_record(ws, mapping: dict) -> None:
    """No.7 就労実績セクションを解析する。"""
    record_cells = find_cells_by_text(ws, '就労実績', 28, 45)
    if not record_cells:
        return
    rec_row = record_cells[0][0]

    yearmonth_cells = find_cells_by_text(ws, '年月', rec_row, rec_row + 3)
    yearmonth_cells = [c for c in yearmonth_cells if c[2].strip() == '年月']
    if yearmonth_cells:
        ym_row = yearmonth_cells[0][0]
        year_labels = [(c.row, c.column) for c in ws[ym_row]
                       if c.value and str(c.value).strip() == '年']
        month_labels = [(c.row, c.column) for c in ws[ym_row]
                        if c.value and str(c.value).strip() == '月']

        year_month_pairs = []
        for yl in year_labels:
            year_input = find_input_before_label(ws, ym_row, yl[1])
            month_input = None
            for ml in month_labels:
                if ml[1] > yl[1]:
                    month_input = find_input_before_label(ws, ym_row, ml[1])
                    break
            year_month_pairs.append((year_input, month_input))

        suffixes = ['1月目', '2月目', '3月目']
        for i, (yi, mi) in enumerate(year_month_pairs[:3]):
            mapping[f'就労実績・{suffixes[i]}・年'] = yi
            mapping[f'就労実績・{suffixes[i]}・月'] = mi

    day_month_cells = find_cells_by_text(ws, '日／月', rec_row, rec_row + 4)
    if day_month_cells:
        dm_row = day_month_cells[0][0]
        day_labels = [(c.row, c.column) for c in ws[dm_row]
                      if c.value and '日／月' in str(c.value)]
        hour_labels = [(c.row, c.column) for c in ws[dm_row]
                       if c.value and '時間／月' in str(c.value)]

        suffixes = ['1月目', '2月目', '3月目']
        for i in range(3):
            if i < len(day_labels):
                mapping[f'就労実績・{suffixes[i]}・日数'] = find_input_before_label(ws, dm_row, day_labels[i][1])
            if i < len(hour_labels):
                mapping[f'就労実績・{suffixes[i]}・時間数'] = find_input_before_label(ws, dm_row, hour_labels[i][1])


def _parse_leave_period(ws, mapping: dict, search_text: str, prefix: str,
                        row_start: int, row_end: int,
                        filter_fn=None) -> None:
    """休業・休暇系セクション（産前産後休業・育児休業等）の期間を解析する共通関数。"""
    cells = find_cells_by_text(ws, search_text, row_start, row_end)
    if filter_fn:
        cells = [c for c in cells if filter_fn(ws, c)]
    if not cells:
        return
    base_row = cells[0][0]
    period_cells = find_cells_by_text(ws, '期間', base_row, base_row + 3)
    if period_cells:
        p = analyze_period_row(ws, period_cells[0][0], period_cells[0][1])
        for k, v in p.items():
            mapping[f'{prefix}・{k}'] = v


def _parse_maternity_leave(ws, mapping: dict) -> None:
    """No.8 産前・産後休業セクションを解析する。"""
    _parse_leave_period(
        ws, mapping, '産前', '産前産後休業', 34, 45,
        filter_fn=lambda ws, c: '産後' in c[2],
    )


def _parse_childcare_leave(ws, mapping: dict) -> None:
    """No.9 育児休業セクションを解析する。"""
    def _filter(ws, c):
        return ('取得' in str(ws.cell(row=c[0], column=9).value or '')
                or '取得' in c[2] or '期間' not in c[2])
    _parse_leave_period(
        ws, mapping, '育児休業', '育児休業', 36, 48,
        filter_fn=_filter,
    )


def _parse_other_leave(ws, mapping: dict) -> None:
    """No.10 産休・育休以外の休業セクションを解析する。"""
    other_leave_cells = find_cells_by_text(ws, '産休・育休以外', 35, 55)
    if not other_leave_cells:
        other_leave_cells = find_cells_by_text(ws, '以外の休業', 35, 55)
    if not other_leave_cells:
        return
    ol_row = other_leave_cells[0][0]

    # その他理由テキスト
    other_reason = find_cells_by_text(ws, 'その他（', ol_row, ol_row + 2)
    if not other_reason:
        other_reason = find_cells_by_text(ws, 'その他(', ol_row, ol_row + 2)
    if other_reason:
        paren_close = [(c.row, c.column) for c in ws[other_reason[0][0]]
                       if c.value and str(c.value).strip() in ('）', ')') and c.column > other_reason[0][1]]
        if paren_close:
            mapping['産休育休以外・その他理由'] = find_input_in_range(
                ws, other_reason[0][0], other_reason[0][1] + 1, paren_close[0][1])
        else:
            for mr in ws.merged_cells.ranges:
                if mr.min_row == other_reason[0][0] and mr.min_col > other_reason[0][1]:
                    mapping['産休育休以外・その他理由'] = cell_addr(mr.min_row, mr.min_col)
                    break

    # 期間
    period_cells = find_cells_by_text(ws, '期間', ol_row, ol_row + 3)
    if period_cells:
        p = analyze_period_row(ws, period_cells[0][0], period_cells[0][1])
        for k, v in p.items():
            mapping[f'産休育休以外・{k}'] = v


def _parse_return_date(ws, mapping: dict) -> None:
    """No.11 復職（予定）年月日セクションを解析する。"""
    return_cells = find_cells_by_text(ws, '復職', 35, 55)
    return_cells = [c for c in return_cells if '年月日' in c[2]]
    if not return_cells:
        return
    ret_row = return_cells[0][0]
    ymd = analyze_ymd_row(ws, ret_row, return_cells[0][1])
    for k, v in ymd.items():
        mapping[f'復職年月日・{k}'] = v


def _parse_short_time_work(ws, mapping: dict) -> None:
    """No.12 育児短時間勤務制度セクションを解析する。"""
    short_cells = find_cells_by_text(ws, '短時間', 35, 55)
    short_cells = [c for c in short_cells if '勤務' in c[2]]
    if not short_cells:
        return
    st_row = short_cells[0][0]

    # 期間
    period_kw = find_cells_by_text(ws, '期間', st_row, st_row + 2)
    if period_kw:
        p = analyze_period_row(ws, period_kw[0][0], period_kw[0][1])
        for k, v in p.items():
            mapping[f'短時間勤務・{k}'] = v

    # 就労時間帯
    time_kw = find_cells_by_text(ws, '就労時間帯', st_row, st_row + 3)
    if not time_kw:
        time_kw = find_cells_by_text(ws, 'シフト', st_row, st_row + 3)
    if time_kw:
        tr = analyze_time_row(ws, time_kw[0][0])
        for k, v in tr.items():
            mapping[f'短時間勤務・就労時間帯・{k}'] = v


def _parse_tanshin_funin(ws, mapping: dict) -> None:
    """No.17 単身赴任期間セクションを解析する。"""
    tanshin_cells = find_cells_by_text(ws, '単身赴任', 40, 60)
    if not tanshin_cells:
        return
    ts_row = tanshin_cells[0][0]
    p = analyze_period_row(ws, ts_row, tanshin_cells[0][1])
    for k, v in p.items():
        mapping[f'単身赴任・{k}'] = v


def _parse_remarks(ws, mapping: dict) -> None:
    """No.18 備考欄セクションを解析する。"""
    remarks_cells = find_cells_by_text(ws, '備考', 40, 60)
    if not remarks_cells:
        return
    rem_row = remarks_cells[0][0]
    rem_label_col = remarks_cells[0][1]
    rem_label_end = _get_label_end_col(ws, rem_row, rem_label_col)
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= rem_row <= mr.max_row and mr.min_col > rem_label_end and (mr.max_col - mr.min_col) >= 5:
            mapping['備考欄'] = cell_addr(mr.min_row, mr.min_col)
            break


def _parse_guardian_section(ws, mapping: dict) -> None:
    """No.19 保護者記載欄（児童名・生年月日・施設名）セクションを解析する。"""
    hogosya_cells = find_cells_by_text(ws, '保護者記載', 40, 65)
    if not hogosya_cells:
        hogosya_cells = find_cells_by_text(ws, '保護者記入', 40, 65)
    if not hogosya_cells:
        return
    h_row = hogosya_cells[0][0]
    logger.debug(f"  保護者記載欄ヘッダー: row={h_row}")

    child_labels = find_cells_by_text(ws, '児童名', h_row, h_row + 12)
    child_idx = 0
    for cl_row, cl_col, _ in child_labels:
        child_idx += 1
        suffix = str(child_idx)
        logger.debug(f"  児童名{suffix}: row={cl_row}, col={cl_col}")

        # 児童名: ラベルの次の行にある同じ列範囲の結合セルが入力セル
        found_child = False
        for mr in ws.merged_cells.ranges:
            if mr.min_row == cl_row + 1 and mr.min_col == cl_col and (mr.max_col - mr.min_col) >= 3:
                mapping[f'保護者記載欄・児童名{suffix}'] = cell_addr(mr.min_row, mr.min_col)
                found_child = True
                break
        # フォールバック: 同行レイアウト
        if not found_child:
            input_addr = find_right_panel_input(ws, cl_row, cl_col)
            if input_addr:
                mapping[f'保護者記載欄・児童名{suffix}'] = input_addr
                found_child = True
                logger.debug(f"    児童名{suffix} (同行フォールバック): {input_addr}")

        # 児童名の入力セルが見つからない場合、この児童のフィールドをスキップ
        if not found_child:
            logger.debug(f"    児童名{suffix}: 入力セルが見つからないためスキップ")
            continue

        # 生年月日
        birth_in_row = [c for c in ws[cl_row]
                       if c.value and '生年' in str(c.value) and c.column > cl_col]
        if birth_in_row:
            ymd = analyze_ymd_row(ws, cl_row + 1, birth_in_row[0].column)
            if not ymd:
                ymd = analyze_ymd_row(ws, cl_row, birth_in_row[0].column)
                if ymd:
                    logger.debug(f"    生年月日{suffix} (同行フォールバック)")
            for k in ('年', '月', '日'):
                if k in ymd:
                    mapping[f'保護者記載欄・生年月日{suffix}・{k}'] = ymd[k]

        # 施設名
        facility_in_row = [c for c in ws[cl_row]
                          if c.value and '施設名' in str(c.value) and c.column > cl_col]
        if facility_in_row:
            fc = facility_in_row[0].column
            found_facility = False
            for mr in ws.merged_cells.ranges:
                if mr.min_row == cl_row + 1 and mr.min_col == fc and (mr.max_col - mr.min_col) >= 3:
                    mapping[f'保護者記載欄・施設名{suffix}'] = cell_addr(mr.min_row, mr.min_col)
                    found_facility = True
                    break
            if not found_facility:
                input_addr = find_right_panel_input(ws, cl_row, fc)
                if input_addr:
                    mapping[f'保護者記載欄・施設名{suffix}'] = input_addr
                    logger.debug(f"    施設名{suffix} (同行フォールバック): {input_addr}")


# ---------------------------------------------------------------------------
# Main analysis: text input fields
# ---------------------------------------------------------------------------

def analyze_certificate(workbook_path: str) -> dict[str, str | None]:
    """就労証明書を解析して全フィールドのマッピングを作成"""
    wb = openpyxl.load_workbook(workbook_path)
    ws = _select_form_sheet(wb)
    mapping = {}

    mapping['自治体'] = extract_municipality_from_content(ws)
    logger.debug(f"自治体: {mapping['自治体']}")

    # 各セクションを順に解析
    _section_parsers = [
        ('宛先', _parse_addressee),
        ('証明日', _parse_certification_date),
        ('事業所情報', _parse_office_info),
        ('No.1 業種', _parse_industry),
        ('No.2 フリガナ・本人氏名・生年月日', _parse_personal_info),
        ('No.3 雇用期間', _parse_employment_period),
        ('No.4 本人就労先事業所', _parse_work_office),
        ('No.5 雇用形態', _parse_employment_type),
        ('No.6 就労時間', _parse_work_time),
        ('No.7 就労実績', _parse_work_record),
        ('No.8 産前産後休業', _parse_maternity_leave),
        ('No.9 育児休業', _parse_childcare_leave),
        ('No.10 産休育休以外', _parse_other_leave),
        ('No.11 復職年月日', _parse_return_date),
        ('No.12 短時間勤務', _parse_short_time_work),
        ('No.17 単身赴任', _parse_tanshin_funin),
        ('No.18 備考欄', _parse_remarks),
        ('No.19 保護者記載欄', _parse_guardian_section),
    ]

    for section_name, parser_fn in _section_parsers:
        logger.debug(f"=== {section_name} ===")
        before_keys = set(mapping.keys())
        parser_fn(ws, mapping)
        new_keys = sorted(set(mapping.keys()) - before_keys)
        if new_keys:
            logger.debug(f"  検出: {new_keys}")
        else:
            logger.debug(f"  検出なし")

    # キーの区切り文字を ・ → _ に変換（自治体キーは除外）
    municipality = mapping.pop('自治体', '不明')
    mapping = {k.replace('・', '_'): v for k, v in mapping.items()}
    mapping['自治体'] = municipality

    logger.debug(f"最終マッピング ({len(mapping)-1} fields): "
                 f"{json.dumps({k:v for k,v in mapping.items() if k != '自治体'}, ensure_ascii=False)}")
    return mapping


# ---------------------------------------------------------------------------
# Comprehensive checkbox detection
# ---------------------------------------------------------------------------

_SECTION_NAME_MAP = {
    '雇用（予定）期間等': '雇用期間',
    '雇用の形態': '雇用形態',
    '就労時間（固定就労の場合）': '就労時間',
    '就労時間（変則就労の場合）': '変則就労',
    '産前･産後休業の取得': '産前産後休業',
    '産前・産後休業の取得': '産前産後休業',
    '育児休業の取得': '育児休業',
    '産休・育休以外の休業の取得': '産休育休以外',
    '復職（予定）年月日': '復職年月日',
    '育児のための短時間勤務制度利用有無': '短時間勤務',
    '保育士等としての勤務実態の有無': '保育士等勤務実態有無',
    '（雇用契約の）満了後の更新の有無': '雇用契約満了後更新有無',
    '入所の内定時における育児休業の短縮の可否': '入所内定時育休短縮可否',
    '育児休業の延長の可否': '育休延長可否',
    '保護者記載欄': '保護者記載欄',
    '保護者記入欄': '保護者記載欄',
}

_KEYWORD_FALLBACKS = [
    ('入所', '育休', '短縮', '入所内定時育休短縮可否'),
    ('育休', '延長', '可否', '育休延長可否'),
]


def _normalize_section_name(raw: str) -> str:
    """セクション名を正規化する（複数行結合、※除外、半角括弧→全角、テキスト名と統一）"""
    lines = [l.strip() for l in raw.strip().split('\n') if not l.strip().startswith('※')]
    text = ''.join(lines)
    text = text.replace('(', '（').replace(')', '）')
    mapped = _SECTION_NAME_MAP.get(text, text)
    if mapped != text:
        return mapped
    for fb in _KEYWORD_FALLBACKS:
        keywords, target = fb[:-1], fb[-1]
        if all(kw in text for kw in keywords):
            return target
    return mapped


def _detect_checkbox_label(ws, row: int, col: int) -> str | None:
    """チェックボックス(□)の右側からラベルテキストを検出する。"""
    label = None
    for next_col in range(col + 1, min(col + 6, ws.max_column + 1)):
        next_val = ws.cell(row=row, column=next_col).value
        if next_val:
            text = str(next_val).strip()
            if text and text != '□':
                label = text
                break
    # フォールバック: 上のセルを確認
    if not label:
        above_val = ws.cell(row=row - 1, column=col).value
        if above_val:
            label = str(above_val).strip()
    if label:
        label = label.rstrip('（(')
        label = label.replace('（第一希望）', '').replace('(第一希望)', '')
        label = label.replace('（第1希望）', '').replace('(第1希望)', '')
    return label


def _detect_row_context(ws, row: int, checkbox_col: int) -> str:
    """チェックボックス行のコンテキスト（合計時間/就労日数/理由）を検出する。"""
    # 「理由」ラベルより右のチェックボックスには「理由_」を付与
    for rc in ws[row]:
        if rc.value and str(rc.value).strip() == '理由' and rc.column < checkbox_col:
            return "理由_"
    # Column I の行コンテキスト
    i_val = ws.cell(row=row, column=9).value
    if i_val and isinstance(i_val, str):
        ctx = i_val.strip().split('\n')[0]
        if ctx in ('合計時間', '就労日数'):
            return f"{ctx}_"
    return ""


def find_all_checkboxes(ws) -> dict[str, str]:
    """すべてのチェックボックス(□)とそのラベルを検出して返す"""
    raw_items: list[tuple[str, str]] = []
    current_section = ""
    _HOGOSYA_SUB_LABELS = {'児童名', '生年月日', '施設名'}

    for row in range(13, ws.max_row + 1):
        # Track section from columns B-H
        for col in range(2, 9):
            c_val = ws.cell(row=row, column=col).value
            if c_val and isinstance(c_val, str):
                text = c_val.strip()
                stripped = text.replace('No.', '').replace('№', '').replace('.', '').strip()
                if len(text) > 1 and not stripped.isdigit():
                    if text in _HOGOSYA_SUB_LABELS and current_section == '保護者記載欄':
                        break
                    current_section = _normalize_section_name(c_val)
                    break

        for cell in ws[row]:
            if not cell.value or str(cell.value).strip() != '□':
                continue

            label = _detect_checkbox_label(ws, row, cell.column)
            if not label:
                continue

            section = current_section if current_section else f"row{row}"
            row_context = _detect_row_context(ws, row, cell.column)
            key = f"{section}_{row_context}{label}"
            raw_items.append((key, cell.coordinate))

    # Deduplicate: keys appearing multiple times get sequential numbers
    key_counts = Counter(k for k, _ in raw_items)
    key_indices: dict[str, int] = {}
    checkboxes: dict[str, str] = {}
    for key, addr in raw_items:
        if key_counts[key] > 1:
            idx = key_indices.get(key, 0) + 1
            key_indices[key] = idx
            checkboxes[f"{key}{idx}"] = addr
        else:
            checkboxes[key] = addr

    logger.debug(f"チェックボックス検出: {len(checkboxes)}件")
    return checkboxes


# ---------------------------------------------------------------------------
# Verification and self-healing
# ---------------------------------------------------------------------------

_EXPECTED_TEXT_SECTIONS = [
    ('宛先', 1),
    ('証明日', 3),
    ('事業所名', 1), ('代表者名', 1), ('所在地', 1),
    ('電話番号', 3), ('担当者名', 1), ('記載者連絡先', 3),
    ('業種', 1),
    ('フリガナ', 1), ('本人氏名', 1), ('生年月日', 3),
    ('雇用期間', 6),
    ('事業所名称', 1), ('事業所住所', 1),
    ('雇用形態', 1),
    ('就労時間_月間', 3), ('就労時間_一月当たり', 1), ('就労時間_一週当たり', 1),
    ('就労時間_平日', 5), ('就労時間_土曜', 5), ('就労時間_日祝', 5),
    ('変則就労_合計時間', 3), ('変則就労_就労日数', 1), ('変則就労_就労時間帯', 5),
    ('就労実績', 12),
    ('産前産後休業', 6),
    ('育児休業', 6),
    ('産休育休以外', 7),
    ('復職年月日', 3),
    ('短時間勤務', 11),
    ('単身赴任', 6),
    ('備考欄', 1),
    ('保護者記載欄_児童名', 3), ('保護者記載欄_生年月日', 9), ('保護者記載欄_施設名', 3),
]

_EXPECTED_CB_SECTIONS = [
    ('業種', 12),
    ('雇用期間', 2),
    ('雇用形態', 6),
    ('就労時間', 8),
    ('変則就労', 1),
    ('産前産後休業', 2),
    ('育児休業', 3),
    ('産休育休以外', 5),
    ('復職年月日', 2),
    ('短時間勤務', 2),
    ('保育士等勤務実態有無', 2),
    ('雇用契約満了後更新有無', 3),
    ('入所内定時育休短縮可否', 2),
    ('育休延長可否', 2),
    ('保護者記載欄', 6),
]


def _repair_text_field_by_scan(ws, mapping: dict, keyword: str, field_key: str,
                               repair_fn) -> int:
    """全行スキャンで特定のテキストフィールドを修復する汎用関数。修復件数を返す。"""
    if field_key in mapping:
        return 0
    logger.debug(f"修復試行: {field_key}")
    for r in range(10, ws.max_row + 1):
        for c in ws[r]:
            if c.value and keyword in str(c.value):
                result = repair_fn(ws, r, c)
                if result is not None:
                    mapping[field_key] = result
                    logger.debug(f"  修復成功: {field_key} = {result}")
                    return 1
    return 0


def _repair_text_fields(ws, mapping: dict) -> int:
    """不足テキストフィールドの修復を試みる。修復件数を返す。"""
    repaired = 0

    # --- 備考欄 ---
    def _repair_remarks(ws, row, cell):
        if '欄' not in str(cell.value):
            return None
        label_end = _get_label_end_col(ws, row, cell.column)
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col > label_end and (mr.max_col - mr.min_col) >= 3:
                return cell_addr(mr.min_row, mr.min_col)
        return None

    repaired += _repair_text_field_by_scan(ws, mapping, '備考', '備考欄', _repair_remarks)

    # --- 単身赴任 ---
    tanshin_keys = [k for k in mapping if k.startswith('単身赴任')]
    if len(tanshin_keys) < 6:
        logger.debug("修復試行: 単身赴任")
        for r in range(10, ws.max_row + 1):
            for c in ws[r]:
                if c.value and '単身赴任' in str(c.value):
                    p = analyze_period_row(ws, r, c.column)
                    for k, v in p.items():
                        key = f'単身赴任_{k}'
                        if key not in mapping:
                            mapping[key] = v
                            repaired += 1
                    if any(k.startswith('単身赴任') for k in mapping):
                        logger.debug(f"  修復成功: 単身赴任 {[k for k in mapping if k.startswith('単身赴任')]}")
                        break
            if len([k for k in mapping if k.startswith('単身赴任')]) >= 6:
                break

    # --- 産休育休以外_その他理由 ---
    def _repair_other_leave_reason(ws, row, cell):
        if '以外' not in str(cell.value) and '育休' not in str(cell.value):
            return None
        for search_row in range(row, min(row + 3, ws.max_row + 1)):
            other_reason = find_cells_by_text(ws, 'その他', search_row, search_row + 1)
            for or_cell in other_reason:
                if '（' in or_cell[2] or '(' in or_cell[2]:
                    paren_close = [(cc.row, cc.column) for cc in ws[or_cell[0]]
                                   if cc.value and str(cc.value).strip() in ('）', ')') and cc.column > or_cell[1]]
                    if paren_close:
                        return find_input_in_range(ws, or_cell[0], or_cell[1] + 1, paren_close[0][1])
        return None

    repaired += _repair_text_field_by_scan(
        ws, mapping, '産休', '産休育休以外_その他理由', _repair_other_leave_reason)

    return repaired


def _repair_checkboxes(ws, checkboxes: dict) -> int:
    """不足チェックボックスの修復を試みる（□以外のマーカーも検出）。修復件数を返す。"""
    repaired = 0
    alt_markers = ['☐', '☑', '✓', '✔']
    existing_addrs = set(checkboxes.values())

    for row in range(13, ws.max_row + 1):
        for cell in ws[row]:
            if not cell.value:
                continue
            val = str(cell.value).strip()
            if val not in alt_markers:
                continue
            if cell.coordinate in existing_addrs:
                continue
            label = _detect_checkbox_label(ws, row, cell.column)
            if not label:
                continue
            key = f"alt_row{row}_{label}"
            checkboxes[key] = cell.coordinate
            existing_addrs.add(cell.coordinate)
            repaired += 1
            logger.debug(f"  修復成功(代替マーカー): {key} = {cell.coordinate}")

    return repaired


def verify_and_repair(ws, mapping: dict, checkboxes: dict) -> None:
    """テキスト/チェックボックスの項目数を検証し、不足分の修復を試みる。"""
    text_count = len(mapping) - (1 if '自治体' in mapping else 0)
    cb_count = len(checkboxes)

    logger.info(f"検証開始: テキスト {text_count}/{EXPECTED_TEXT_COUNT}, "
                f"チェックボックス {cb_count}/{EXPECTED_CHECKBOX_COUNT}")

    for prefix, expected in _EXPECTED_TEXT_SECTIONS:
        actual = len([k for k in mapping if k.startswith(prefix) and k != '自治体'])
        if actual < expected:
            logger.warning(f"  テキスト不足: {prefix} ({actual}/{expected})")

    if text_count < EXPECTED_TEXT_COUNT:
        logger.info(f"テキスト修復開始 (不足: {EXPECTED_TEXT_COUNT - text_count}件)")
        text_repaired = _repair_text_fields(ws, mapping)
        new_text_count = len(mapping) - (1 if '自治体' in mapping else 0)
        logger.info(f"テキスト修復完了: {text_repaired}件追加 → {new_text_count}/{EXPECTED_TEXT_COUNT}")
    else:
        logger.info("テキストフィールド: 修復不要")

    if cb_count < EXPECTED_CHECKBOX_COUNT:
        logger.info(f"チェックボックス修復開始 (不足: {EXPECTED_CHECKBOX_COUNT - cb_count}件)")
        cb_repaired = _repair_checkboxes(ws, checkboxes)
        logger.info(f"チェックボックス修復完了: {cb_repaired}件追加 → {len(checkboxes)}/{EXPECTED_CHECKBOX_COUNT}")
    else:
        logger.info("チェックボックス: 修復不要")

    final_text = len(mapping) - (1 if '自治体' in mapping else 0)
    final_cb = len(checkboxes)
    if final_text < EXPECTED_TEXT_COUNT:
        missing = EXPECTED_TEXT_COUNT - final_text
        logger.warning(f"最終結果: テキスト {missing}件不足 ({final_text}/{EXPECTED_TEXT_COUNT})")
        for prefix, expected in _EXPECTED_TEXT_SECTIONS:
            actual = len([k for k in mapping if k.startswith(prefix) and k != '自治体'])
            if actual < expected:
                logger.warning(f"  未解決: {prefix} ({actual}/{expected})")
    else:
        logger.info(f"最終結果: テキスト OK ({final_text}/{EXPECTED_TEXT_COUNT})")

    if final_cb < EXPECTED_CHECKBOX_COUNT:
        logger.warning(f"最終結果: チェックボックス {EXPECTED_CHECKBOX_COUNT - final_cb}件不足 "
                       f"({final_cb}/{EXPECTED_CHECKBOX_COUNT})")
    else:
        logger.info(f"最終結果: チェックボックス OK ({final_cb}/{EXPECTED_CHECKBOX_COUNT})")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    workbook_path = sys.argv[1]
    flags = sys.argv[2:]
    json_mode = '--json' in flags
    log_file = str(Path(workbook_path).parent / 'analyze_shuroushomei_debug.log')
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(levelname)s: %(message)s',
        handlers=[
            logging.StreamHandler(sys.stderr),
            logging.FileHandler(log_file, mode='w', encoding='utf-8'),
        ],
    )

    if not Path(workbook_path).exists():
        if json_mode:
            print(json.dumps({"error": f"File not found: {workbook_path}"}, ensure_ascii=False))
            sys.exit(1)
        print(f'Error: File not found: {workbook_path}')
        sys.exit(1)

    mapping = analyze_certificate(workbook_path)
    wb = openpyxl.load_workbook(workbook_path)
    ws = _select_form_sheet(wb)
    checkboxes = find_all_checkboxes(ws)

    verify_and_repair(ws, mapping, checkboxes)

    municipality = mapping.pop('自治体', '不明')

    sheet_name = ws.title

    if json_mode:
        output = {
            "municipality": municipality,
            "sheet_name": sheet_name,
            "text": mapping,
            "checkbox": checkboxes,
        }
        print(json.dumps(output, ensure_ascii=False, indent=2))
        return

    print(f'Analyzing: {Path(workbook_path).name}')
    print(f'  Sheet: {sheet_name}')
    print(f'  Municipality: {municipality}')
    print(f'  Found {len(mapping)} text input fields')
    for k, v in sorted(mapping.items()):
        print(f'    {k}: {v}')

    print(f'\n  Found {len(checkboxes)} checkboxes')
    for k, v in sorted(checkboxes.items()):
        print(f'    {k}: {v}')


if __name__ == '__main__':
    main()
